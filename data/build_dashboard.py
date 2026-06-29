"""
Genera el archivo Excel del dashboard My English Spot.
Tres pestañas: Search Console, GBP, Instagram.
"""
import base64
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
import csv, io

# ── Colores de marca ─────────────────────────────────────────────────────────
BLUE       = "FF65C6F4"   # --blue
BLUE_DEEP  = "FF2EA7DA"   # --blue-deep
ORANGE     = "FFFFA51F"   # --orange
INK        = "FF1F2330"   # --ink
CREAM      = "FFFAF8F3"   # --cream
WHITE      = "FFFFFFFF"
LIGHT_BLUE = "FFE8F7FD"
LIGHT_ORA  = "FFFEF3DC"
GRAY_LINE  = "FFEBE6D8"
GREEN_SOFT = "FFD6F5D6"

def hdr(cell, bg, fg=INK, bold=True, size=10):
    cell.font = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def body(cell, bg=WHITE, fg=INK, bold=False, align="left"):
    cell.font = Font(name="Calibri", bold=bold, color=fg, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")

def thin_border():
    s = Side(style="thin", color=GRAY_LINE)
    return Border(left=s, right=s, top=s, bottom=s)

def apply_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── PESTAÑA 1: Search Console ─────────────────────────────────────────────────
def build_search_console(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    # Fila 1: título del dashboard
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "MY ENGLISH SPOT — Search Console"
    t.font = Font(name="Calibri", bold=True, color=WHITE, size=13)
    t.fill = PatternFill("solid", fgColor=BLUE_DEEP)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    # Fila 2: metadata del export
    ws.merge_cells("A2:E2")
    m = ws["A2"]
    m.value = "Período: últimos 3 meses · Tipo de búsqueda: Web · Exportado desde Google Search Console"
    m.font = Font(name="Calibri", italic=True, color="FF666666", size=9)
    m.fill = PatternFill("solid", fgColor=CREAM)
    m.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    # ── BLOQUE A: Rendimiento diario (Gráfico.csv) ────────────────────────────
    ws["A3"] = "📅  Rendimiento diario"
    ws["A3"].font = Font(name="Calibri", bold=True, color=BLUE_DEEP, size=10)
    ws["A3"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.merge_cells("A3:E3")
    ws.row_dimensions[3].height = 20

    headers_daily = ["Fecha", "Clics", "Impresiones", "CTR", "Posición media"]
    for col, h in enumerate(headers_daily, 1):
        c = ws.cell(row=4, column=col, value=h)
        hdr(c, BLUE_DEEP, WHITE)
    ws.row_dimensions[4].height = 22

    daily_data = [
        ("2026-05-17", 0, 0, None, None),
        ("2026-05-18", 0, 1, "0%", 13),
        ("2026-05-19", 0, 0, None, None),
        ("2026-05-20", 3, 4, "75%", 1.2),
        ("2026-05-21", 2, 3, "66.67%", 1.0),
        ("2026-05-22", 0, 0, None, None),
        ("2026-05-23", 0, 0, None, None),
        ("2026-05-24", 3, 10, "30%", 2.4),
        ("2026-05-25", 0, 7, "0%", 6.6),
        ("2026-05-26", 1, 10, "10%", 4.9),
        ("2026-05-27", 0, 1, "0%", 8.0),
        ("2026-05-28", 0, 6, "0%", 3.7),
        ("2026-05-29", 4, 13, "30.77%", 5.4),
    ]
    start_row = 5
    for i, row_data in enumerate(daily_data):
        r = start_row + i
        bg = CREAM if i % 2 == 0 else WHITE
        vals = [row_data[0], row_data[1], row_data[2],
                row_data[3] if row_data[3] else "—",
                row_data[4] if row_data[4] else "—"]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            body(c, bg, align="center" if col > 1 else "left")

    last_daily = start_row + len(daily_data) - 1
    apply_border(ws, 4, last_daily, 1, 5)

    # ── BLOQUE B: Consultas principales ──────────────────────────────────────
    blk2 = last_daily + 2
    ws.cell(row=blk2, column=1, value="🔍  Consultas principales")
    ws.cell(row=blk2, column=1).font = Font(name="Calibri", bold=True,
                                            color=ORANGE, size=10)
    ws.cell(row=blk2, column=1).fill = PatternFill("solid", fgColor=LIGHT_ORA)
    ws.merge_cells(f"A{blk2}:E{blk2}")

    hdr_row = blk2 + 1
    for col, h in enumerate(["Consulta", "Clics", "Impresiones", "CTR", "Posición media"], 1):
        c = ws.cell(row=hdr_row, column=col, value=h)
        hdr(c, ORANGE, WHITE)

    queries = [
        ("clases particulares ingles vigo", 0, 2, "0%", 13),
        ("my english school",               0, 1, "0%", 1),
        ("open english",                    0, 1, "0%", 3),
        ("english school",                  0, 1, "0%", 5),
        ("english",                         0, 1, "0%", 9),
        ("clases particulares de ingles en vigo", 0, 1, "0%", 10),
        ("6 minutes english",               0, 1, "0%", 12),
        ("academia inglés vigo",            0, 1, "0%", 14),
        ("clases de inglés vigo",           0, 1, "0%", 14),
    ]
    for i, (q, clics, impr, ctr, pos) in enumerate(queries):
        r = hdr_row + 1 + i
        bg = CREAM if i % 2 == 0 else WHITE
        for col, v in enumerate([q, clics, impr, ctr, pos], 1):
            c = ws.cell(row=r, column=col, value=v)
            body(c, bg, align="center" if col > 1 else "left")
    last_q = hdr_row + len(queries)
    apply_border(ws, hdr_row, last_q, 1, 5)

    # ── BLOQUE C: Páginas / Dispositivos / Países (lado a lado) ──────────────
    blk3 = last_q + 2
    ws.cell(row=blk3, column=1, value="📄  Páginas principales")
    ws.cell(row=blk3, column=1).font = Font(name="Calibri", bold=True,
                                            color=BLUE_DEEP, size=10)
    ws.cell(row=blk3, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.merge_cells(f"A{blk3}:E{blk3}")

    hdr3 = blk3 + 1
    for col, h in enumerate(["Página", "Clics", "Impresiones", "CTR", "Posición"], 1):
        c = ws.cell(row=hdr3, column=col, value=h)
        hdr(c, BLUE_DEEP, WHITE)
    pages = [("https://www.myenglishspotclasses.com/", 13, 55, "23.64%", 4.36)]
    for i, row in enumerate(pages):
        r = hdr3 + 1 + i
        for col, v in enumerate(row, 1):
            c = ws.cell(row=r, column=col, value=v)
            body(c, CREAM if i % 2 == 0 else WHITE,
                 align="left" if col == 1 else "center")
    apply_border(ws, hdr3, hdr3 + len(pages), 1, 5)

    blk4 = hdr3 + len(pages) + 2
    ws.cell(row=blk4, column=1, value="💻  Dispositivos")
    ws.cell(row=blk4, column=1).font = Font(name="Calibri", bold=True,
                                            color=ORANGE, size=10)
    ws.cell(row=blk4, column=1).fill = PatternFill("solid", fgColor=LIGHT_ORA)
    ws.merge_cells(f"A{blk4}:E{blk4}")
    hdr4 = blk4 + 1
    for col, h in enumerate(["Dispositivo", "Clics", "Impresiones", "CTR", "Posición"], 1):
        c = ws.cell(row=hdr4, column=col, value=h)
        hdr(c, ORANGE, WHITE)
    devices = [("Ordenador", 8, 28, "28.57%", 5.18),
               ("Móviles",   5, 27, "18.52%", 3.52)]
    for i, row in enumerate(devices):
        r = hdr4 + 1 + i
        for col, v in enumerate(row, 1):
            c = ws.cell(row=r, column=col, value=v)
            body(c, CREAM if i % 2 == 0 else WHITE,
                 align="left" if col == 1 else "center")
    apply_border(ws, hdr4, hdr4 + len(devices), 1, 5)

    blk5 = hdr4 + len(devices) + 2
    ws.cell(row=blk5, column=1, value="🌍  Países")
    ws.cell(row=blk5, column=1).font = Font(name="Calibri", bold=True,
                                            color=BLUE_DEEP, size=10)
    ws.cell(row=blk5, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.merge_cells(f"A{blk5}:E{blk5}")
    hdr5 = blk5 + 1
    for col, h in enumerate(["País", "Clics", "Impresiones", "CTR", "Posición"], 1):
        c = ws.cell(row=hdr5, column=col, value=h)
        hdr(c, BLUE_DEEP, WHITE)
    countries = [("España", 13, 55, "23.64%", 4.36)]
    for i, row in enumerate(countries):
        r = hdr5 + 1 + i
        for col, v in enumerate(row, 1):
            c = ws.cell(row=r, column=col, value=v)
            body(c, CREAM if i % 2 == 0 else WHITE,
                 align="left" if col == 1 else "center")
    apply_border(ws, hdr5, hdr5 + len(countries), 1, 5)

    set_col_widths(ws, [42, 10, 14, 10, 14])


# ── PESTAÑA 2: GBP ───────────────────────────────────────────────────────────
def build_gbp(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "MY ENGLISH SPOT — Google Business Profile"
    t.font = Font(name="Calibri", bold=True, color=WHITE, size=13)
    t.fill = PatternFill("solid", fgColor=ORANGE)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:C2")
    m = ws["A2"]
    m.value = "Período: dic 2025 – may 2026  (6 meses)"
    m.font = Font(name="Calibri", italic=True, color="FF666666", size=9)
    m.fill = PatternFill("solid", fgColor=CREAM)
    m.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    data_blocks = [
        ("🔭  Visibilidad", BLUE_DEEP, LIGHT_BLUE, [
            ("Vistas del Perfil de Empresa", "1.053", "vistas"),
            ("Búsquedas que mostraron el perfil", "41", "búsquedas"),
        ]),
        ("👆  Interacciones", ORANGE, LIGHT_ORA, [
            ("Interacciones del Perfil de Empresa", "273", "interacciones"),
            ("Solicitudes Cómo llegar", "238", "solicitudes"),
            ("Clics al sitio web", "30", "clics"),
        ]),
        ("📱  Plataforma (vistas)", BLUE_DEEP, LIGHT_BLUE, [
            ("Búsqueda Google (móvil)", "620", "59%"),
            ("Google Maps (móvil)", "177", "17%"),
            ("Búsqueda Google (ordenador)", "175", "17%"),
            ("Google Maps (ordenador)", "81", "8%"),
        ]),
        ("🔎  Términos de búsqueda", ORANGE, LIGHT_ORA, [
            ("spot", "23", "apariciones"),
            ("english", "18", "apariciones"),
            ("6 minutes english", "<15", "apariciones"),
            ("academia ingles vigo", "<15", "apariciones"),
            ("academia ingles vigo florida", "<15", "apariciones"),
        ]),
    ]

    current_row = 3
    for (section_title, hdr_color, bg_color, rows) in data_blocks:
        # Sección header
        ws.cell(row=current_row, column=1, value=section_title)
        ws.cell(row=current_row, column=1).font = Font(
            name="Calibri", bold=True,
            color=hdr_color if hdr_color != LIGHT_BLUE else BLUE_DEEP,
            size=10)
        ws.cell(row=current_row, column=1).fill = PatternFill("solid", fgColor=bg_color)
        ws.merge_cells(f"A{current_row}:C{current_row}")
        current_row += 1

        # Col headers
        for col, h in enumerate(["Métrica", "Valor", "Unidad"], 1):
            c = ws.cell(row=current_row, column=col, value=h)
            hdr(c, hdr_color, WHITE)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Datos
        for i, (metrica, valor, unidad) in enumerate(rows):
            bg = CREAM if i % 2 == 0 else WHITE
            ws.cell(row=current_row, column=1, value=metrica)
            body(ws.cell(row=current_row, column=1), bg)
            ws.cell(row=current_row, column=2, value=valor)
            body(ws.cell(row=current_row, column=2), bg, bold=True, align="center")
            ws.cell(row=current_row, column=3, value=unidad)
            body(ws.cell(row=current_row, column=3), bg, align="center")
            current_row += 1

        apply_border(ws, current_row - len(rows) - 1, current_row - 1, 1, 3)
        current_row += 1  # espacio entre bloques

    set_col_widths(ws, [40, 14, 16])


# ── PESTAÑA 3: Instagram ──────────────────────────────────────────────────────
def build_instagram(ws):
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "MY ENGLISH SPOT — Instagram"
    t.font = Font(name="Calibri", bold=True, color=WHITE, size=13)
    t.fill = PatternFill("solid", fgColor="FF833AB4")   # morado Instagram
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:G2")
    m = ws["A2"]
    m.value = "Datos pendientes de exportar desde Instagram Insights"
    m.font = Font(name="Calibri", italic=True, color="FF888888", size=9)
    m.fill = PatternFill("solid", fgColor=CREAM)
    m.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    headers = ["Semana", "Seguidores", "Alcance", "Impresiones",
               "Interacciones", "Posts publicados", "Notas"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        hdr(c, "FF833AB4", WHITE)
    ws.row_dimensions[3].height = 22

    apply_border(ws, 3, 3, 1, len(headers))
    set_col_widths(ws, [16, 13, 11, 14, 16, 18, 30])


# ── BUILD ─────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# Renombrar la hoja por defecto y añadir las demás
wb.active.title = "Search Console"
ws_sc  = wb.active
ws_gbp = wb.create_sheet("GBP")
ws_ig  = wb.create_sheet("Instagram")

# Colores de las pestañas
ws_sc.sheet_properties.tabColor  = "2EA7DA"
ws_gbp.sheet_properties.tabColor = "FFA51F"
ws_ig.sheet_properties.tabColor  = "833AB4"

build_search_console(ws_sc)
build_gbp(ws_gbp)
build_instagram(ws_ig)

output_path = "/root/my-english-spot/data/dashboard.xlsx"
wb.save(output_path)

# Retornar base64 para el MCP upload
with open(output_path, "rb") as f:
    b64 = base64.b64encode(f.read()).decode()

print(b64)
